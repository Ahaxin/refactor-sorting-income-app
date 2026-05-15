"""
Generates a 4-sheet Excel workbook:
  1. Good Life Report  — daily income and per-person salary breakdown
  2. Tianyuan Report   — same structure
  3. Schedule          — master (employee × day) calendar
  4. Summary           — monthly totals, formula verification, seed/phantom metadata
"""

import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.config import GOOD_LIFE, TIANYUAN, OUTPUT_FILE, SE_SALARY_UNIT
from src.models.employee import SelfEmployedEmployee, CompanyEmployedEmployee
from src.models.company import Company

logger = logging.getLogger(__name__)

CLR_HEADER = "4472C4"
CLR_TOTAL_ROW = "D9E1F2"
CLR_GL = "E2EFDA"
CLR_TY = "FFF2CC"
CLR_CE_ONLY = "FFFFD7"
CLR_WARN = "FF0000"


def generate_report(
    se_workers: list[SelfEmployedEmployee],
    ce_workers: list[CompanyEmployedEmployee],
    companies: dict[str, Company],
    seed: int = 0,
    output_path: str | None = None,
) -> None:
    path = output_path if output_path is not None else OUTPUT_FILE
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    _write_company_sheet(wb, GOOD_LIFE, "Good Life", se_workers, ce_workers, companies)
    _write_company_sheet(wb, TIANYUAN, "Tianyuan", se_workers, ce_workers, companies)
    _write_schedule_sheet(wb, se_workers, ce_workers, companies)
    _write_summary_sheet(wb, se_workers, ce_workers, companies, seed=seed)

    wb.save(path)
    logger.info(f"Report saved to: {path}")


def _header_font() -> Font:
    return Font(bold=True, color="FFFFFF")


def _header_fill(colour: str = CLR_HEADER) -> PatternFill:
    return PatternFill("solid", fgColor=colour)


def _thin_border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_header(cell, label: str, colour: str = CLR_HEADER):
    cell.value = label
    cell.font = _header_font()
    cell.fill = _header_fill(colour)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _apply_cell(cell, value, fill_colour: str | None = None, bold: bool = False, red: bool = False):
    cell.value = value
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin_border()
    if fill_colour:
        cell.fill = PatternFill("solid", fgColor=fill_colour)
    if bold:
        cell.font = Font(bold=True)
    if red:
        cell.font = Font(color=CLR_WARN, bold=True)


def _write_company_sheet(
    wb: Workbook,
    company_key: str,
    sheet_title: str,
    se_workers: list[SelfEmployedEmployee],
    ce_workers: list[CompanyEmployedEmployee],
    companies: dict[str, Company],
) -> None:
    company = companies[company_key]
    ws = wb.create_sheet(title=sheet_title)

    ce_here = [w for w in ce_workers if any(c == company_key for (c, _) in w.schedule)]
    se_here = [w for w in se_workers if any(c == company_key for (c, _) in w.schedule)]

    headers = ["Day", "Income", "Cleaned Income"]
    headers += [f"CE: {w.name}" for w in ce_here]
    headers += ["CE Total", "SE Target"]
    headers += [f"SE: {w.name}" for w in se_here]
    headers += ["SE Total", "Formula Check\n(SE/0.4+CE)"]

    for col_idx, h in enumerate(headers, start=1):
        _apply_header(ws.cell(row=1, column=col_idx), h)
    ws.freeze_panes = "A2"

    monthly_ce: dict[str, int] = {w.name: 0 for w in ce_here}
    monthly_se: dict[str, int] = {w.name: 0 for w in se_here}
    monthly_income = 0
    monthly_cleaned = 0
    monthly_se_total = 0
    monthly_ce_total = 0

    for row_idx, day in enumerate(company.days, start=2):
        dl = company.get_day(day)
        row_colour = CLR_CE_ONLY if dl.is_full_ce_absorption else None

        col = 1
        _apply_cell(ws.cell(row=row_idx, column=col), day, fill_colour=row_colour)
        col += 1
        _apply_cell(ws.cell(row=row_idx, column=col), dl.raw_income, fill_colour=row_colour)
        col += 1
        _apply_cell(ws.cell(row=row_idx, column=col), dl.cleaned_income, fill_colour=row_colour)
        col += 1

        for w in ce_here:
            val = dl.ce_salaries.get(w.name, 0)
            _apply_cell(ws.cell(row=row_idx, column=col), val if val else "", fill_colour=row_colour)
            monthly_ce[w.name] += val
            col += 1

        ce_total = dl.ce_total
        _apply_cell(ws.cell(row=row_idx, column=col), ce_total, fill_colour=row_colour, bold=True)
        monthly_ce_total += ce_total
        col += 1

        _apply_cell(ws.cell(row=row_idx, column=col), dl.se_day_target, fill_colour=row_colour)
        col += 1

        for w in se_here:
            val = dl.se_salaries.get(w.name, 0)
            _apply_cell(ws.cell(row=row_idx, column=col), val if val else "", fill_colour=row_colour)
            monthly_se[w.name] += val
            col += 1

        se_total = dl.se_total
        _apply_cell(ws.cell(row=row_idx, column=col), se_total, fill_colour=row_colour, bold=True)
        monthly_se_total += se_total
        col += 1

        if dl.is_full_ce_absorption:
            _apply_cell(ws.cell(row=row_idx, column=col), "CE-FULL", fill_colour=CLR_CE_ONLY)
        else:
            fval = round(dl.formula_check)
            deviation = abs(fval - dl.cleaned_income)
            cell = ws.cell(row=row_idx, column=col)
            _apply_cell(cell, fval, fill_colour=row_colour)
            if deviation > SE_SALARY_UNIT:
                cell.font = Font(color=CLR_WARN, bold=True)

        monthly_income += dl.raw_income
        monthly_cleaned += dl.cleaned_income

    total_row = len(company.days) + 2
    _apply_cell(ws.cell(row=total_row, column=1), "MONTHLY TOTAL", fill_colour="D9E1F2", bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    col = 4
    for w in ce_here:
        _apply_cell(ws.cell(row=total_row, column=col), monthly_ce[w.name], fill_colour="D9E1F2", bold=True)
        col += 1
    _apply_cell(ws.cell(row=total_row, column=col), monthly_ce_total, fill_colour="D9E1F2", bold=True)
    col += 2
    for w in se_here:
        actual = monthly_se[w.name]
        target = w.salary
        is_short = actual != target
        _apply_cell(ws.cell(row=total_row, column=col), actual, fill_colour="D9E1F2", bold=True, red=is_short)
        col += 1
    _apply_cell(ws.cell(row=total_row, column=col), monthly_se_total, fill_colour="D9E1F2", bold=True)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13
    ws.row_dimensions[1].height = 40


def _write_schedule_sheet(
    wb: Workbook,
    se_workers: list[SelfEmployedEmployee],
    ce_workers: list[CompanyEmployedEmployee],
    companies: dict[str, Company],
) -> None:
    ws = wb.create_sheet(title="Schedule")
    all_days = sorted(next(iter(companies.values())).days)
    all_workers = [*se_workers, *ce_workers]

    _apply_header(ws.cell(row=1, column=1), "Employee")
    for col_idx, day in enumerate(all_days, start=2):
        _apply_header(ws.cell(row=1, column=col_idx), str(day))
    ws.freeze_panes = "B2"

    for row_idx, worker in enumerate(all_workers, start=2):
        name_cell = ws.cell(row=row_idx, column=1)
        name_cell.value = worker.name
        name_cell.font = Font(bold=True)
        name_cell.border = _thin_border()
        name_cell.alignment = Alignment(horizontal="left", vertical="center")
        for col_idx, day in enumerate(all_days, start=2):
            company_on_day = worker.company_on_day(day)
            cell = ws.cell(row=row_idx, column=col_idx)
            if company_on_day == GOOD_LIFE:
                _apply_cell(cell, "GL", fill_colour=CLR_GL)
            elif company_on_day == TIANYUAN:
                _apply_cell(cell, "TY", fill_colour=CLR_TY)
            else:
                _apply_cell(cell, "")

    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(all_days) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 5
    ws.row_dimensions[1].height = 25


def _write_summary_sheet(
    wb: Workbook,
    se_workers: list[SelfEmployedEmployee],
    ce_workers: list[CompanyEmployedEmployee],
    companies: dict[str, Company],
    seed: int = 0,
) -> None:
    ws = wb.create_sheet(title="Summary")
    row = 1

    # Run metadata
    _apply_header(ws.cell(row=row, column=1), "Run Metadata", "1F497D")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    _apply_cell(ws.cell(row=row, column=1), "Random Seed", bold=True)
    _apply_cell(ws.cell(row=row, column=2), seed)
    row += 2

    # SE/CE salary table
    headers = ["Name", "Type", "Monthly Target / Cap", "Actual Earned", "Gap", "Status"]
    for col_idx, h in enumerate(headers, start=1):
        _apply_header(ws.cell(row=row, column=col_idx), h)
    row += 1

    for worker in se_workers:
        target = worker.salary
        actual = worker.actual_monthly_salary
        gap = actual - target
        status = "OK" if gap == 0 else "SHORTFALL"
        is_short = status == "SHORTFALL"
        for col_idx, val in enumerate([worker.name, "SE", target, actual, gap, status], start=1):
            _apply_cell(ws.cell(row=row, column=col_idx), val, red=is_short)
        row += 1

    for worker in ce_workers:
        cap = worker.salary
        actual = worker.actual_monthly_salary
        gap = actual - cap
        status = "OK" if actual <= cap else "OVER CAP"
        for col_idx, val in enumerate([worker.name, "CE", cap, actual, gap, status], start=1):
            _apply_cell(ws.cell(row=row, column=col_idx), val)
        row += 1

    row += 1

    # Per-day formula verification
    _apply_header(ws.cell(row=row, column=1), "Formula Verification per Day (SE/0.4 + CE = I_clean)", "1F497D")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    day_headers = ["Company", "Day", "I_clean", "SE Target", "SE Total", "CE Total",
                   "Formula\n(SE/0.4+CE)", "Status"]
    for col_idx, h in enumerate(day_headers, start=1):
        _apply_header(ws.cell(row=row, column=col_idx), h)
    row += 1

    for company in companies.values():
        for day in company.days:
            dl = company.get_day(day)
            if dl.is_full_ce_absorption:
                vals = [company.name, day, dl.cleaned_income, 0, 0, dl.ce_total, "—", "CE-FULL"]
                is_bad = False
            else:
                fval = round(dl.formula_check)
                dev = fval - dl.cleaned_income
                status = "OK" if abs(dev) <= SE_SALARY_UNIT else f"DEVIATION {dev:+d}"
                vals = [company.name, day, dl.cleaned_income, dl.se_day_target,
                        dl.se_total, dl.ce_total, fval, status]
                is_bad = abs(dev) > SE_SALARY_UNIT
            for col_idx, v in enumerate(vals, start=1):
                _apply_cell(ws.cell(row=row, column=col_idx), v, red=is_bad)
            row += 1

    for col_idx, width in enumerate([14, 5, 12, 10, 10, 10, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
